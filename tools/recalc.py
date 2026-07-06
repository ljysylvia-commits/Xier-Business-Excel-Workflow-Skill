#!/usr/bin/env python3
"""recalc —— Excel 公式重算 + 错误计数（LibreOffice headless，跨场景共享工具）

用法:
  python3 recalc.py <文件.xlsx> [--timeout 30]

行为:
  1) 用 soffice --headless 重新加载并另存该 xlsx（触发公式重算）
  2) 用 openpyxl(data_only=True) 扫描全表，统计 #REF! / #DIV/0! / #VALUE! / #NAME? 数量
  3) stdout 打印 JSON: {"status": "success|error", "total_errors": N, "detail": {...}}
退出码: 0=重算完成且 total_errors==0 / 1=有公式错误 / 2=环境或执行失败
注意: 必须在 x14 扩展注入【之前】调用本工具（LibreOffice 保存会吞掉 x14 扩展）。
"""
import argparse, json, os, shutil, subprocess, sys, tempfile

ERRS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--timeout", type=int, default=30)
    a = ap.parse_args()
    path = os.path.abspath(a.file)
    if not os.path.isfile(path):
        print(json.dumps({"status": "error", "msg": f"文件不存在: {path}"})); sys.exit(2)
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        print(json.dumps({"status": "error", "msg": "未找到 soffice/libreoffice"})); sys.exit(2)
    with tempfile.TemporaryDirectory() as td:
        try:
            r = subprocess.run(
                [soffice, "--headless", "--norestore", "--convert-to", "xlsx", "--outdir", td, path],
                capture_output=True, text=True, timeout=a.timeout)
        except subprocess.TimeoutExpired:
            print(json.dumps({"status": "error", "msg": f"重算超时(>{a.timeout}s)"})); sys.exit(2)
        if r.returncode != 0:
            print(json.dumps({"status": "error", "msg": r.stderr.strip()[:300]})); sys.exit(2)
        out = os.path.join(td, os.path.basename(path))
        if not os.path.exists(out):
            print(json.dumps({"status": "error", "msg": "soffice 未产出转换文件"})); sys.exit(2)
        shutil.copy2(out, path)  # 用重算后的文件覆盖原文件
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(json.dumps({"status": "error", "msg": "缺少 openpyxl"})); sys.exit(2)
    wb = load_workbook(path, data_only=True)
    detail = {e: 0 for e in ERRS}
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and v in detail:
                    detail[v] += 1
    total = sum(detail.values())
    print(json.dumps({"status": "success", "total_errors": total,
                      "detail": {k: v for k, v in detail.items() if v}}, ensure_ascii=False))
    sys.exit(0 if total == 0 else 1)


if __name__ == "__main__":
    main()
