#!/usr/bin/env python3
"""drift_check - deterministic structure drift checks for registered data cleansing suites.

The tool reads a data cleansing suite directory, inspects its cleansing.yaml fingerprint,
and checks whether a new input file still fits the registered structure.

Exit codes:
  0 = pass, no detected drift
  1 = fail, required structure is missing or changed
  2 = partial, input is runnable only after user review
"""
import argparse
import csv
import json
import os
import sys
from pathlib import Path

try:
    import yaml
except ImportError:  # pragma: no cover - exercised in clean environments
    yaml = None


def load_yaml(path):
    if yaml is None:
        raise SystemExit("PyYAML is required for drift_check: pip install pyyaml")
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def status_from_findings(errors, warnings):
    if errors:
        return "fail"
    if warnings:
        return "partial"
    return "pass"


def check_csv(path, fingerprint):
    errors, warnings, checks = [], [], []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        columns = reader.fieldnames or []
    required = fingerprint.get("required_columns") or fingerprint.get("expected_headers") or []
    missing = [col for col in required if col not in columns]
    extra = [col for col in columns if col not in required] if required else []
    checks.append({"check": "csv_columns", "expected": required, "actual": columns, "missing": missing, "extra": extra})
    if missing:
        errors.append(f"missing required columns: {missing}")
    if extra and fingerprint.get("allow_extra_columns") is False:
        warnings.append(f"extra columns require review: {extra}")
    return errors, warnings, checks


def dynamic_sheet_matches(sheetnames, dynamic_sheets):
    matches = []
    misses = []
    for spec in dynamic_sheets or []:
        contains = spec.get("name_contains") or []
        hit = []
        for sheet in sheetnames:
            if all(token in sheet for token in contains):
                hit.append(sheet)
        item = {"spec": spec, "matches": hit}
        matches.append(item)
        if not hit:
            misses.append(spec)
    return matches, misses


def check_xlsx(path, fingerprint):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("openpyxl is required to inspect xlsx inputs: pip install openpyxl")

    wb = load_workbook(path, read_only=True, data_only=True)
    sheetnames = list(wb.sheetnames)
    errors, warnings, checks = [], [], []

    required_sheets = fingerprint.get("required_sheets") or []
    dynamic_sheets = fingerprint.get("dynamic_sheets") or []
    missing_required = [name for name in required_sheets if name not in sheetnames]
    dynamic_matches, missing_dynamic = dynamic_sheet_matches(sheetnames, dynamic_sheets)
    known = set(required_sheets)
    for item in dynamic_matches:
        known.update(item["matches"])
    unknown = [name for name in sheetnames if name not in known]

    checks.append({
        "check": "sheets",
        "sheetnames": sheetnames,
        "required_sheets": required_sheets,
        "missing_required_sheets": missing_required,
        "dynamic_matches": dynamic_matches,
        "unknown_sheets": unknown,
    })
    if missing_required:
        errors.append(f"missing required sheets: {missing_required}")
    if missing_dynamic:
        errors.append(f"missing dynamic sheet matches: {missing_dynamic}")
    if unknown and not fingerprint.get("allow_unknown_sheets", False):
        warnings.append(f"unknown sheets require review: {unknown}")

    expected_headers = fingerprint.get("expected_headers") or []
    header_row = int(fingerprint.get("header_row") or 1)
    data_sheets = fingerprint.get("data_sheets") or [
        name for name in required_sheets if name in sheetnames
    ]
    if expected_headers:
        for sheet in data_sheets:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            actual = [ws.cell(header_row, idx).value for idx in range(1, len(expected_headers) + 1)]
            ok = actual == expected_headers
            checks.append({
                "check": "headers",
                "sheet": sheet,
                "header_row": header_row,
                "expected": expected_headers,
                "actual": actual,
                "match": ok,
            })
            if not ok:
                errors.append(f"header drift in {sheet}: {actual}")
    return errors, warnings, checks


def run_check(suite_dir, input_path):
    suite_dir = Path(suite_dir).resolve()
    input_path = Path(input_path).resolve()
    yaml_path = suite_dir / "cleansing.yaml"
    calibers_path = suite_dir / "CALIBERS.md"
    data = load_yaml(yaml_path)
    understand = data.get("understand") or {}
    mode = understand.get("mode", "fixed")
    fingerprint = understand.get("fingerprint") or {}

    errors, warnings, checks = [], [], []
    if not calibers_path.exists():
        warnings.append("CALIBERS.md is missing; structure baseline cannot be reviewed")
    else:
        text = calibers_path.read_text(encoding="utf-8", errors="replace")
        checks.append({"check": "calibers_structure_baseline", "present": "结构依赖基线" in text})
        if "结构依赖基线" not in text:
            warnings.append("CALIBERS.md does not contain a structure baseline section")

    if mode != "fixed":
        warnings.append(f"mode={mode}; run profile detection or user confirmation before reuse")
        status = status_from_findings(errors, warnings)
        return {
            "status": status,
            "mode": mode,
            "cleansing_id": data.get("cleansing_id") or suite_dir.name,
            "suite_dir": str(suite_dir),
            "input": str(input_path),
            "checks": checks,
            "errors": errors,
            "warnings": warnings,
        }

    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        e, w, c = check_csv(input_path, fingerprint)
    elif suffix in {".xlsx", ".xlsm"}:
        e, w, c = check_xlsx(input_path, fingerprint)
    else:
        e, w, c = [f"unsupported input type: {suffix}"], [], []
    errors.extend(e)
    warnings.extend(w)
    checks.extend(c)
    status = status_from_findings(errors, warnings)
    return {
        "status": status,
        "mode": mode,
        "cleansing_id": data.get("cleansing_id") or suite_dir.name,
        "suite_dir": str(suite_dir),
        "input": str(input_path),
        "checks": checks,
        "errors": errors,
        "warnings": warnings,
    }


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--suite-dir", required=True)
    parser.add_argument("--input", required=True)
    parser.add_argument("--out", help="Optional JSON report path")
    parser.add_argument("--json", action="store_true", help="Print full JSON to stdout")
    args = parser.parse_args()

    result = run_check(args.suite_dir, args.input)
    if args.out:
        Path(args.out).parent.mkdir(parents=True, exist_ok=True)
        Path(args.out).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(f"drift_check: {result['status']} ({result['cleansing_id']})")
        for item in result["errors"] + result["warnings"]:
            print(f"- {item}")

    if result["status"] == "pass":
        sys.exit(0)
    if result["status"] == "partial":
        sys.exit(2)
    sys.exit(1)


if __name__ == "__main__":
    main()
